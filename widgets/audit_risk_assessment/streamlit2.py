# widget/rag_audit_assistant/app.py
"""
RAG Audit Assistant widget: wrapped in `render()` so it mounts inside your host app.
No set_page_config here (the page wrapper will handle that).
"""
import streamlit as st
import pandas as pd
import tempfile
import os
import io
import faiss
import numpy as np
import pdfplumber
import re
import requests
import tiktoken
from typing import Dict, List, Tuple
import time
import math
import openpyxl
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from pydantic import BaseModel, Field, model_validator, ValidationError
import json

# ----- Models (unchanged) -----
class InherentRisk(BaseModel):
    risk_type: str
    Fraud_Risk_Factor: str = Field(..., pattern="^(Yes|No)$")
    Internal_Controls: str
    Likelihood: str = Field(..., pattern="^(High|Low)$")
    Likelihood_Explanation: str
    Material_Quantitative_Impact: str = Field(..., pattern="^(High|Low)$")
    Impact_Explanation: str
    Conclusion: str
    SR: str = Field(..., pattern="^(SR|No SR)$")

    @model_validator(mode="after")
    def coerce_significant_risk(self):
        if self.Likelihood == "High" and self.Material_Quantitative_Impact == "High":
            object.__setattr__(self, "SR", "SR")
        else:
            object.__setattr__(self, "SR", "No SR")
        return self

# ---- One-time secrets/init (cached across reruns) ----
@st.cache_resource(show_spinner=False)
def _init_endpoints() -> Dict[str, str]:
    return {
        "LLM_ENDPOINT": st.secrets["AZURE_API_ENDPOINT"],
        "LLM_API_KEY": st.secrets["AZURE_API_KEY"],
        "EMBED_ENDPOINT": st.secrets["EMBEDDING_ENDPOINT"],
    }

ENC = tiktoken.get_encoding("cl100k_base")

# ---- Helpers (unchanged logic) ----
def clean_text(raw: str) -> str:
    text = re.sub(r'[\x00-\x1F\x7F]+', ' ', raw)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def load_main_body(file) -> str:
    pieces = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            x0,y0,x1,y1 = page.bbox
            crop = (x0+50, y0+50, x1-50, y1-50)
            txt = page.within_bbox(crop).extract_text() or ""
            pieces.append(txt)
    full = " ".join(pieces)
    return clean_text(full)

def chunk_text_words(text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
    words = text.split()
    stride = chunk_size - chunk_overlap
    chunks = []
    for i in range(0, len(words), stride):
        seg = words[i:i+chunk_size]
        if seg:
            chunks.append(" ".join(seg))
    return chunks

def embed_texts(texts: List[str], EMBED_ENDPOINT: str, EMBED_API_KEY: str) -> np.ndarray:
    headers = {"Content-Type": "application/json", "api-key": EMBED_API_KEY}
    r = requests.post(EMBED_ENDPOINT, headers=headers, json={"input": texts})
    r.raise_for_status()
    embs = [d.get("embedding") for d in r.json().get("data", [])]
    arr = np.array(embs, dtype=np.float32)
    faiss.normalize_L2(arr)
    return arr

def index_documents(docs: Dict[str,str], chunk_size: int, chunk_overlap: int, EMBED_ENDPOINT: str, EMBED_API_KEY: str):
    all_chunks, meta = [], []
    for name, text in docs.items():
        chunks = chunk_text_words(text, chunk_size, chunk_overlap)
        for idx,c in enumerate(chunks):
            all_chunks.append(c)
            meta.append((name, idx))
    if not all_chunks:
        return None, [], []
    embeddings = embed_texts(all_chunks, EMBED_ENDPOINT, EMBED_API_KEY)
    index = faiss.IndexFlatIP(embeddings.shape[1])
    index.add(embeddings)
    return index, meta, all_chunks

def retrieve_chunks(query: str, index, meta, chunks, k: int, EMBED_ENDPOINT: str, EMBED_API_KEY: str):
    if index is None:
        return []
    q_emb = embed_texts([query], EMBED_ENDPOINT, EMBED_API_KEY)[0]
    D,I = index.search(np.array([q_emb]), k)
    return [(meta[i][0], chunks[i]) for i in I[0]]

def get_llm_response(prompt: str, context: str, LLM_ENDPOINT: str, LLM_API_KEY: str) -> tuple[str, str]:
    headers = {"Content-Type": "application/json", "api-key": LLM_API_KEY}
    system_message = (
        "You are a world-class corporate analysis assistant for an expert audit team. "
        "Use the context below to answer due diligence questions. Use the internet to answer any questions you aren't aware of the answers to. \n\n"
        f"Context:\n{context}\n\n"
        "Respond in natural, flowing English paragraphs. Do not use any markdown syntax. "
        "Cite your sources in square brackets like [1], [2], etc. at the relevant point in the response. "
        "At the end, include a list of the sources used, each on a new line, prefixed with the corresponding number. Be very specific about the part of the website, and the subpage used to take the info from. Don't just list the main page of the webpage."
        "Example: [1] https://example.com/example_sub_directory"
    )
    messages = [
        {"role": "system", "content": system_message},
        {"role": "user", "content": prompt}
    ]
    r = requests.post(LLM_ENDPOINT, headers=headers, json={"messages": messages})
    r.raise_for_status()
    time.sleep(3)
    response_text = r.json()["choices"][0]["message"]["content"].strip()
    if "\n[1]" in response_text:
        split_index = response_text.find("\n[1]")
        main_answer = response_text[:split_index].strip()
        sources = response_text[split_index:].strip()
    else:
        main_answer = response_text
        sources = ""
    return main_answer, sources

def _coerce_high_low(value: str | None) -> str:
    if not value:
        return "Low"
    value = value.strip().lower()
    return "High" if "high" in value else "Low"

def _coerce_yes_no(value: str | None) -> str:
    if not value:
        return "No"
    value = value.strip().lower()
    return "Yes" if "yes" in value else "No"

def parse_risks_response(raw: str) -> List[InherentRisk]:
    fence_pattern = re.compile(r"^```(?:json)?\s*(.*)\s*```$", re.DOTALL)
    m = fence_pattern.match(raw.strip())
    if m:
        raw = m.group(1)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"Response was not valid JSON: {e}\n\nRaw text:\n{raw}")
    if not isinstance(data, list):
        raise ValueError(f"Expected a JSON array, got {type(data)}")
    risks: List[InherentRisk] = []
    for idx, item in enumerate(data):
        mapped = {
            "risk_type":                    item.get("risk_type"),
            "Fraud_Risk_Factor":            _coerce_yes_no(item.get("Fraud Risk Factor?")),
            "Internal_Controls":            item.get("Internal Controls", "Not provided"),
            "Likelihood":                   _coerce_high_low(item.get("Likelihood")),
            "Likelihood_Explanation":       item.get("Likelihood Explanation", "Not provided"),
            "Material_Quantitative_Impact": _coerce_high_low(item.get("Material Quantitative Impact?")),
            "Impact_Explanation":           item.get("Impact Explanation", "Not provided"),
            "Conclusion":                   item.get("Conclusion", "Not provided"),
            "SR":                           "No SR",
        }
        try:
            risks.append(InherentRisk(**mapped))
        except ValidationError as ve:
            print(f"Warning: Item {idx} failed validation and was skipped:\n{ve}\nRaw: {item}")
    return risks

def get_structure_llm_response(prompt, context, company_name, book_year, LLM_ENDPOINT: str, LLM_API_KEY: str):
    schema_instruction = """
    When you answer, return only a JSON array of objects, each matching exactly this schema:
    [
      {
        "risk_type":           "string",
        "Fraud Risk Factor?":  "Yes"|"No",
        "Internal Controls":   "string",
        "Likelihood":          "High"|"Low",
        "Likelihood Explanation":"string",
        "Material Quantitative Impact?":"High"|"Low",
        "Impact Explanation":  "string",
        "Conclusion":          "string",
        "SR?":                 "SR"|"No SR"
      }
    ]
    """
    headers = {"Content-Type":"application/json","api-key":LLM_API_KEY}
    messages=[{"role":"system","content":(
            "You are an expert audit assistant. "
            f"Use the provided context to identify *sources of inherent risk*, within the enterprise {company_name} in the year {book_year}"
            "and fill in each field per item of risk that you identify. Keep each field short, max 5-10 words\n"
            + schema_instruction +"\n\n"+ context
            )},
              {"role":"user","content": f"Use the provided context to identify *sources of inherent risk*, within the enterprise {company_name} in the year {book_year}"
            "and fill in each field per item of risk that you identify. Keep each field short, max 5-10 words\n"}]
    r = requests.post(LLM_ENDPOINT, headers=headers, json={"messages":messages})
    r.raise_for_status()
    time.sleep(3)
    return r.json()["choices"][0]["message"]["content"].strip()

# ----- The widget UI -----
def render() -> None:
    eps = _init_endpoints()
    LLM_ENDPOINT = eps["LLM_ENDPOINT"]
    LLM_API_KEY = eps["LLM_API_KEY"]
    EMBED_ENDPOINT = eps["EMBED_ENDPOINT"]
    EMBED_API_KEY  = LLM_API_KEY

    # Title + settings
    st.write("Upload your public & client PDFs and the Input GPT.xlsx. Click **Run** to retrieve relevant chunks & generate answers.")

    with st.sidebar:
        st.header("Settings")
        chunk_size    = st.number_input("Chunk size (words)",   min_value=50,   max_value=1000, value=200, step=50)
        chunk_overlap = st.number_input("Chunk overlap (words)",min_value=0,    max_value=chunk_size-1, value=50, step=10)
        top_k = st.number_input("Chunks per query (k)", min_value=1, max_value=20, value=5, step=1)

    public_files = st.file_uploader("Public PDFs", type="pdf", accept_multiple_files=True, key="pubs")
    client_files = st.file_uploader("Client PDFs", type="pdf", accept_multiple_files=True, key="clients")
    company_name  = st.text_input("Company Name")
    audit_year = st.text_input("Year for Audit")

    excel_path = Path("Input GPT.xlsx")
    if excel_path.exists() and "excel_bytes" not in st.session_state:
        st.session_state.excel_bytes = excel_path.read_bytes()

    risks_template_path = Path("empty template.xlsx")
    if risks_template_path.exists() and "risks_excel_bytes" not in st.session_state:
        st.session_state.risks_excel_bytes = risks_template_path.read_bytes()

    st.session_state.setdefault("pipeline_ran", False)

    # ---- Outputs already computed ----
    if st.session_state.pipeline_ran:
        st.success("✅ Retrieval complete! Download your updated workbooks below.")
        st.download_button(
            "Download 1300 workbook",
            data=st.session_state.out1,
            file_name="1300_only.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.download_button(
            "Download Memo workbook",
            data=st.session_state.out2,
            file_name="Memo_only.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        if st.button("Clear Results and Rerun Retrieval"):
            st.session_state.pipeline_ran = False
            st.session_state.pop("out1", None)
            st.session_state.pop("out2", None)

    # ---- Otherwise, show Run ----
    else:
        if st.button("Run Retrieval & Generate Answers"):
            if "excel_bytes" not in st.session_state:
                st.error("Please upload the Input GPT.xlsx file.")
                st.stop()
            if "risks_excel_bytes" not in st.session_state:
                st.error("Please include the empty template.xlsx file.")
                st.stop()

            original_bytes = st.session_state.excel_bytes
            original_bytes_risks =  st.session_state.risks_excel_bytes
            df1 = pd.read_excel(BytesIO(original_bytes), sheet_name="1300")
            df2 = pd.read_excel(BytesIO(original_bytes_risks), sheet_name="Memo")

            # Prepare the second‐sheet columns
            cols = ['Fraud Risk Factor?','Internal Controls','Likelihood','Likelihood Explanation',
                    'Material Quantitative Impact?','Impact Explanation','Conclusion','SR?']
            prompts = {  # kept for future use if you prompt per-column
                'Fraud Risk Factor?':       "The Fraud Risk Factor of the above risk type. Answer ONLY with Yes or No.",
                'Internal Controls':        "What are the internal controls within the company against this type of risk. Answer with a maximum of 10 words.",
                'Likelihood':               "Based on public and previous sources. What is the likelihood of this type of risk occurring. Answer ONLY with High or Low.",
                'Likelihood Explanation':   "Based on public and previous sources. Only include justification, max 15 words.",
                'Material Quantitative Impact?': "Based on public and previous sources. Estimated impact of this type of risk. Answer ONLY with High or Low.",
                'Impact Explanation':       "Based on public and previous sources. Only include justification, max 15 words.",
                'Conclusion':               "Explain if there is significant risk or if further discussion is needed. Max 10-15 words.",
                'SR?':                      "Answer only with 'SR' or 'No SR'."
            }
            for c in cols:
                if c not in df2.columns:
                    df2[c] = ""

            total = len(df1)
            pb = st.progress(0.0)
            pt = st.empty()
            processed = 0

            pub_docs = {f.name: load_main_body(f) for f in public_files}
            cli_docs = {f.name: load_main_body(f) for f in client_files}
            pub_idx, pub_meta, pub_chunks = index_documents(pub_docs, chunk_size, chunk_overlap, EMBED_ENDPOINT, EMBED_API_KEY)
            cli_idx, cli_meta, cli_chunks = index_documents(cli_docs, chunk_size, chunk_overlap, EMBED_ENDPOINT, EMBED_API_KEY)

            df1[['Generated answer','Sources']] = ""
            for i, row in df1.iterrows():
                if isinstance(row["#"], float) and math.isnan(row["#"]):
                    continue
                prompt = row['Question']
                example = row['Best practice answer']

                hits_pub = retrieve_chunks(prompt, pub_idx, pub_meta, pub_chunks, top_k, EMBED_ENDPOINT, EMBED_API_KEY)
                hits_cli = retrieve_chunks(prompt, cli_idx, cli_meta, cli_chunks, top_k, EMBED_ENDPOINT, EMBED_API_KEY)
                ctx = "PUBLIC CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_pub)
                ctx += "\n\nCLIENT CONTEXT:\n" + "\n".join(f"[{s}] {t}" for s,t in hits_cli)
                ctx += "\n\nFORMAT EXAMPLE:\n Q: {prompt}\n A: {example}"
                ctx += "\nDO NOT RESPOND WITH MARKDOWN"

                ans, sources = get_llm_response(prompt, ctx, LLM_ENDPOINT, LLM_API_KEY)
                df1.at[i, 'Generated answer'] = ans
                df1.at[i, 'Sources'] = sources
                processed += 1
                pb.progress(processed / total)
                pt.text(f"Answered {processed} of {total} questions")

            # Insert company and year
            df2.iloc[2, 2] = company_name
            df2.iloc[4, 2] = audit_year
            prompt =  (
                f"Use the provided context to identify *sources of inherent risk*, within the enterprise "
                f"{company_name} in the year {audit_year} and fill in each field per item of risk that you identify. "
                "Keep each field short, max 5-10 words\n"
            )
            ctx = "PUBLIC CONTEXT:\n" + "\n".join(f"[{f}]: {pub_docs[f]}" for f in pub_docs.keys())
            ctx += "\n\nCLIENT CONTEXT:\n" + "\n".join(f"[{f}] {cli_docs[f]}" for f in cli_docs.keys())
            raw = get_structure_llm_response(prompt, ctx, company_name, audit_year, LLM_ENDPOINT, LLM_API_KEY)
            risks_list = parse_risks_response(raw)

            start_excel_row = 31
            start_idx = start_excel_row - 2

            needed = start_idx + len(risks_list)
            if len(df2) < needed:
                n_extra = needed - len(df2)
                blank = pd.DataFrame([[None]*len(df2.columns)] * n_extra, columns=df2.columns)
                df2 = pd.concat([df2, blank], ignore_index=True)

            for i, risk in enumerate(risks_list):
                row = start_idx + i
                df2.iat[row, 0] = i + 1
                df2.iat[row, 1] = risk.risk_type
                field_map = {
                    'Fraud Risk Factor?':            risk.Fraud_Risk_Factor,
                    'Internal Controls':             risk.Internal_Controls,
                    'Likelihood':                    risk.Likelihood,
                    'Likelihood Explanation':        risk.Likelihood_Explanation,
                    'Material Quantitative Impact?': risk.Material_Quantitative_Impact,
                    'Impact Explanation':            risk.Impact_Explanation,
                    'Conclusion':                    risk.Conclusion,
                    'SR?':                           risk.SR,
                }
                for j, key in enumerate(field_map.keys()):
                    df2.iat[row, j+2] = field_map[key]

            # Workbook 1 – preserve formatting in “1300” sheet
            wb1 = load_workbook(BytesIO(st.session_state.excel_bytes))
            ws1 = wb1["1300"]
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, max_col=len(df1.columns)):
                for cell in row:
                    try: cell.value = None
                    except AttributeError: pass
            for r, row_vals in enumerate(df1.values, start=2):
                for c, val in enumerate(row_vals, start=1):
                    try: ws1.cell(row=r, column=c).value = val
                    except AttributeError: pass
            out1 = BytesIO(); wb1.save(out1); out1.seek(0)

            # Workbook 2 – preserve formatting in “Memo” sheet
            wb2 = load_workbook(BytesIO(st.session_state.risks_excel_bytes))
            ws2 = wb2["Memo"]
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(df2.columns)):
                for cell in row:
                    try: cell.value = None
                    except AttributeError: pass
            for r, row_vals in enumerate(df2.values, start=2):
                for c, val in enumerate(row_vals, start=1):
                    try: ws2.cell(row=r, column=c).value = val
                    except AttributeError: pass
            out2 = BytesIO(); wb2.save(out2); out2.seek(0)

            st.session_state.out1 = out1
            st.session_state.out2 = out2
            st.session_state.pipeline_ran = True

            st.success("All done!")
            st.download_button("Download 1300 workbook", data=out1, file_name="1300.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.download_button("Download Overview of risks workbook", data=out2, file_name="Overview_of_risks.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
